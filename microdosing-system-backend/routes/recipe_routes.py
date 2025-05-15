from flask import Blueprint, request, jsonify, send_file
from extensions import db
from models.recipe import Recipe, RecipeMaterial, RecipeSchema
from models.production import ProductionOrder
from models.user import User
from sqlalchemy.exc import IntegrityError
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from barcode import Code128
from barcode.writer import ImageWriter
from PIL import Image as PILImage
import io, os, tempfile
from werkzeug.exceptions import BadRequest
import logging
import time
from sqlalchemy.orm import joinedload
from extensions import socketio


recipe_bp = Blueprint("recipe", __name__)
logging.basicConfig(level=logging.DEBUG)

@recipe_bp.route("/recipes/export/barcodes", methods=["GET"])
def export_recipes_excel_with_barcodes():
    start_time = time.time()
    try:
        recipes = Recipe.query.with_entities(Recipe.name, Recipe.code, Recipe.barcode_id).all()
        wb = Workbook()
        ws = wb.active
        ws.title = "Recipes with Barcodes"
        ws.append(["Name", "Code", "Barcode ID", "Scannable Barcode"])

        row_number = 2
        for recipe in recipes:
            name, code, barcode_id = recipe
            if barcode_id:
                try:
                    temp_dir = tempfile.gettempdir()
                    filename = f"{barcode_id}"
                    filepath = os.path.join(temp_dir, f"{filename}.png")
                    code128 = Code128(barcode_id, writer=ImageWriter())
                    code128.save(filepath)
                    image = PILImage.open(filepath)
                    image = image.resize((200, 60))
                    image.save(filepath)

                    ws.cell(row=row_number, column=1, value=name)
                    ws.cell(row=row_number, column=2, value=code)
                    ws.cell(row=row_number, column=3, value=barcode_id)

                    img = ExcelImage(filepath)
                    img.width = 150
                    img.height = 50
                    ws.add_image(img, f"D{row_number}")

                    os.remove(filepath)
                    row_number += 1

                except Exception as e:
                    logging.error(f"Failed to generate barcode for {barcode_id}: {e}")
                    ws.cell(row=row_number, column=1, value=name)
                    ws.cell(row=row_number, column=2, value=code)
                    ws.cell(row=row_number, column=3, value=barcode_id)
                    row_number += 1

        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)

        end_time = time.time()
        execution_time_ms = round((end_time - start_time) * 1000, 2)

        response = send_file(
            stream,
            download_name="recipes_with_barcodes.xlsx",
            as_attachment=True,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response.headers['X-Execution-Time-MS'] = str(execution_time_ms)
        return response

    except Exception as e:
        logging.error(f"Error exporting barcodes: {str(e)}")
        return jsonify({"error": str(e)}), 500

@recipe_bp.route("/recipes", methods=["GET"])
def get_recipes():
    start_time = time.time()
    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 20, type=int)
    pagination = Recipe.query.paginate(page=page, per_page=per_page, error_out=False)
    recipes = pagination.items

    result = [
        {
            "recipe_id": r.recipe_id,
            "name": r.name,
            "code": r.code,
            "description": r.description,
            "version": r.version,
            "status": r.status,
            "created_by": r.created_by,
            "created_at": r.created_at,
            "no_of_materials": r.no_of_materials,
        } for r in recipes
    ]
    end_time = time.time()
    execution_time_ms = round((end_time - start_time) * 1000, 2)

    return jsonify({
        "recipes": result,
        "total": pagination.total,
        "page": pagination.page,
        "pages": pagination.pages,
        "per_page": pagination.per_page,
        "execution_time_ms": execution_time_ms
    })

@recipe_bp.route("/recipes", methods=["POST"])
def create_recipe():
    start_time = time.time()
    data = request.get_json()
    required_fields = ["name", "code", "version", "created_by"]
    for field in required_fields:
        if not data.get(field):
            return jsonify({"error": f"'{field}' is required."}), 400

    status = data.get("status", "Unreleased")
    valid_statuses = ["Released", "Unreleased"]
    if status not in valid_statuses:
        return jsonify({"error": f"Invalid status value: {status}"}), 400

    user = db.session.get(User, data["created_by"])
    if not user:
        return jsonify({"error": "User not found."}), 400

    new_recipe = Recipe(
        name=data["name"],
        code=data["code"],
        description=data.get("description"),
        version=data["version"],
        status=status,
        created_by=data["created_by"],
        barcode_id=data.get("barcode_id"),
        no_of_materials=data.get("no_of_materials")
    )

    db.session.add(new_recipe)
    try:
        db.session.commit()
    except IntegrityError as e:
        db.session.rollback()
        if "Duplicate entry" in str(e.orig):
            return jsonify({"error": "Duplicate entry: code or barcode_id already exists."}), 400
        return jsonify({"error": "Database error occurred."}), 500
    except Exception as e:
        db.session.rollback()
        logging.error(f"Error creating recipe: {str(e)}")
        return jsonify({"error": str(e)}), 500

    end_time = time.time()
    execution_time_ms = round((end_time - start_time) * 1000, 2)
    return jsonify({"message": "Recipe created successfully!", "execution_time_ms": execution_time_ms}), 201

@recipe_bp.route("/recipes/<int:recipe_id>", methods=["GET"])
def get_recipe(recipe_id):
    start_time = time.time()
    recipe = Recipe.query.get(recipe_id)
    if not recipe:
        return jsonify({"error": "Recipe not found"}), 404

    result = {
        "recipe_id": recipe.recipe_id,
        "name": recipe.name,
        "code": recipe.code,
        "description": recipe.description,
        "version": recipe.version,
        "status": recipe.status,
        "created_by": recipe.created_by,
        "created_at": recipe.created_at,
        "no_of_materials": recipe.no_of_materials
    }
    end_time = time.time()
    execution_time_ms = round((end_time - start_time) * 1000, 2)

    return jsonify({**result, "execution_time_ms": execution_time_ms})

@recipe_bp.route("/recipes/<int:recipe_id>", methods=["PUT"])
def update_recipe(recipe_id):
    start_time = time.time()
    try:
        recipe = Recipe.query.get(recipe_id)
        if not recipe:
            return jsonify({"message": "Recipe not found"}), 404

        data = request.get_json()
        recipe.name = data.get("name", recipe.name)
        recipe.code = data.get("code", recipe.code)
        recipe.description = data.get("description", recipe.description)
        recipe.version = data.get("version", recipe.version)
        recipe.status = data.get("status", recipe.status)
        recipe.no_of_materials = data.get("no_of_materials", recipe.no_of_materials)
        if "sequence" in data:
            recipe.sequence = data["sequence"]

        db.session.commit()
        end_time = time.time()
        return jsonify({"message": "Recipe updated successfully", "execution_time_ms": round((end_time - start_time) * 1000, 2)}), 200

    except Exception as e:
        db.session.rollback()
        logging.error(f"Error updating recipe {recipe_id}: {str(e)}")
        return jsonify({"message": "An error occurred while updating the recipe."}), 500

@recipe_bp.route("/recipes/<int:recipe_id>", methods=["DELETE"])
def delete_recipe(recipe_id):
    start_time = time.time()
    try:
        db.session.query(ProductionOrder).filter(ProductionOrder.recipe_id == recipe_id).delete(synchronize_session=False)
        db.session.query(RecipeMaterial).filter(RecipeMaterial.recipe_id == recipe_id).delete(synchronize_session=False)
        recipe = Recipe.query.get(recipe_id)
        if not recipe:
            return jsonify({"message": "Recipe not found"}), 404

        db.session.delete(recipe)
        db.session.commit()
        end_time = time.time()
        return jsonify({"message": "Recipe and related data deleted successfully", "execution_time_ms": round((end_time - start_time) * 1000, 2)}), 200

    except IntegrityError as e:
        db.session.rollback()
        return jsonify({"message": "Integrity error, check related records for consistency"}), 400
    except Exception as e:
        db.session.rollback()
        return jsonify({"message": "An error occurred while deleting the recipe"}), 500

@recipe_bp.route("/recipe_materials", methods=["POST"])
def create_or_update_recipe_material():
    start_time = time.time()
    try:
        data = request.get_json()
        if not data:
            raise BadRequest("No input data provided.")

        recipe_id = data.get("recipe_id")
        material_id = data.get("material_id")
        set_point = data.get("set_point")
        status = data.get("status")
        bucket_id = data.get("bucket_id")
        use_scale = data.get("use_scale", False)

        if use_scale:
            from models.scale import ScaleClient
            scale_client = ScaleClient()
            actual = scale_client.get_net_weight()
            if actual is None:
                return jsonify({"error": "Failed to read weight from scale."}), 500
        else:
            actual = data.get("actual")

        if not recipe_id or not material_id or set_point is None or not status:
            raise BadRequest("Missing required fields.")
        if actual is None:
            raise BadRequest("Actual weight is required.")
        if not isinstance(recipe_id, int) or not isinstance(material_id, int):
            raise BadRequest("recipe_id and material_id must be integers.")
        if not isinstance(set_point, (int, float)) or not isinstance(actual, (int, float)):
            raise BadRequest("set_point and actual must be numeric.")

        if bucket_id is not None:
            from models import StorageBucket
            if not StorageBucket.query.get(bucket_id):
                raise BadRequest("Invalid bucket_id.")

        margin = 0.0 if set_point == 0 else round(((float(set_point) - float(actual)) / float(set_point)) * 100, 2)

        existing_recipe_material = RecipeMaterial.query.filter_by(recipe_id=recipe_id, material_id=material_id).first()

        if existing_recipe_material:
            existing_recipe_material.set_point = set_point
            existing_recipe_material.actual = actual
            existing_recipe_material.margin = margin
            existing_recipe_material.status = status
            existing_recipe_material.bucket_id = bucket_id
            db.session.commit()

            end_time = time.time()
            socketio.emit("recipe_material_updated", {
    "recipe_id": recipe_id,
    "material_id": material_id,
    "actual": actual,
    "set_point": set_point,
    "status": status,
    "bucket_id": bucket_id,
    "margin": margin,
    "type": "update"
})

            return jsonify({
    "message": "Recipe material updated successfully!",
    "execution_time_ms": round((end_time - start_time) * 1000, 2)
}), 200

        else:
            new_recipe_material = RecipeMaterial(
                recipe_id=recipe_id,
                material_id=material_id,
                set_point=set_point,
                actual=actual,
                margin=margin,
                status=status,
                bucket_id=bucket_id
            )
            db.session.add(new_recipe_material)
            db.session.commit()

            end_time = time.time()
            socketio.emit("recipe_material_created", {
    "recipe_id": recipe_id,
    "material_id": material_id,
    "actual": actual,
    "set_point": set_point,
    "status": status,
    "bucket_id": bucket_id,
    "margin": margin,
    "type": "create"
})

            return jsonify({
    "message": "Recipe material created successfully!",
    "execution_time_ms": round((end_time - start_time) * 1000, 2)
}), 201


    except BadRequest as e:
        logging.error(f"Bad request: {e.description}")
        return jsonify({"error": e.description}), 400
    except Exception as e:
        logging.error(f"Unexpected error: {str(e)}")
        return jsonify({"error": "An unexpected error occurred."}), 500




@recipe_bp.route("/recipe_materials/active", methods=["GET"])
def get_active_recipe_material():
    try:
        # Step 1: Get all 'released' recipe materials with material info
        released_materials = (
            db.session.query(RecipeMaterial)
            .options(joinedload(RecipeMaterial.material))
            .filter(RecipeMaterial.status == "pending")
            .order_by(RecipeMaterial.recipe_id, RecipeMaterial.recipe_material_id)
            .all()
        )

        if not released_materials:
            return jsonify({"message": "No released recipe materials available."}), 200

        # Step 2: Group materials by recipe_id
        recipe_materials_map = {}
        for rm in released_materials:
            recipe_materials_map.setdefault(rm.recipe_id, []).append(rm)

        # Step 3: Check for first unapproved recipe and return its first released material
        for recipe_id in sorted(recipe_materials_map.keys()):
            materials = recipe_materials_map[recipe_id]

            approved_exists = (
                db.session.query(RecipeMaterial)
                .filter(RecipeMaterial.recipe_id == recipe_id, RecipeMaterial.status == "approved")
                .first()
            )

            if not approved_exists:
                current_material = materials[0]

                # Use relationship to access name and barcode
                material = current_material.material
                recipe = current_material.recipe
                if material:
                    response_data = {
                        "recipe_name": recipe.name,
                        "material_id": current_material.material_id,
                        "material_name": material.title,
                        "barcode": material.barcode_id,
                        "set_point": current_material.set_point,
                        "actual": current_material.actual,
                        "margin": current_material.margin,
                        "status": current_material.status,
                        "bucket_id": current_material.bucket_id
                    }

                    # Emit WebSocket event to notify clients of the active recipe material
                    emit('new_active_recipe_material', response_data, broadcast=True)

                    return jsonify(response_data), 200

        return jsonify({"message": "No active material ready for display."}), 200

    except Exception as e:
        logging.error(f"Unexpected error in fetching active recipe material: {str(e)}")
        return jsonify({"error": "Internal server error."}), 500

@recipe_bp.route("/recipe_materials/weigh-and-update", methods=["POST"])
def weigh_and_update_material():
    try:
        # Step 1: Mocked weight (replace with real scale reading later)
        actual_weight_kg = 15.25  # Static test value in kg

        # Step 2: Get the active recipe material
        released_materials = (
            db.session.query(RecipeMaterial)
            .options(joinedload(RecipeMaterial.material))
            .filter(RecipeMaterial.status == "pending")
            .order_by(RecipeMaterial.recipe_id, RecipeMaterial.recipe_material_id)
            .all()
        )

        if not released_materials:
            return jsonify({"success": False, "message": "No released recipe materials available."}), 200

        recipe_materials_map = {}
        for rm in released_materials:
            recipe_materials_map.setdefault(rm.recipe_id, []).append(rm)

        for recipe_id in sorted(recipe_materials_map.keys()):
            materials = recipe_materials_map[recipe_id]
            approved_exists = (
                db.session.query(RecipeMaterial)
                .filter(RecipeMaterial.recipe_id == recipe_id, RecipeMaterial.status == "approved")
                .first()
            )
            if not approved_exists:
                current_material = materials[0]

                # Step 3: Update actual and margin
                current_material.actual = actual_weight_kg
                current_material.margin = round(actual_weight_kg - current_material.set_point, 3)
                #current_material.status = "approved"  # Optional

                db.session.commit()

                return jsonify({
                    "success": True,
                    "message": "Mock weight updated successfully.",
                    "data": {
                        "material_id": current_material.material_id,
                        "material_name": current_material.material.title,
                        "actual": current_material.actual,
                        "set_point": current_material.set_point,
                        "margin": current_material.margin
                    }
                }), 200

        return jsonify({"success": False, "message": "No active recipe material ready for update."}), 200

    except Exception as e:
        logging.error(f"Error in weigh_and_update_material: {str(e)}")
        return jsonify({"success": False, "message": "Internal server error."}), 500


@recipe_bp.route("/recipe_materials", methods=["GET"])
def get_recipe_materials():
    start_time = time.time()
    materials = RecipeMaterial.query.all()
    result = [
        {
            "recipe_material_id": mat.recipe_material_id,
            "recipe_id": mat.recipe_id,
            "material_id": mat.material_id,
            "set_point": str(mat.set_point) if mat.set_point is not None else None,
            "actual": str(mat.actual) if mat.actual is not None else None,
            "margin": str(mat.margin) if mat.margin is not None else None
        }
        for mat in materials
    ]
    end_time = time.time()
    return jsonify({"materials": result, "execution_time_ms": round((end_time - start_time) * 1000, 2)})

@recipe_bp.route("/recipe_materials/<int:recipe_material_id>", methods=["PUT"])
def update_recipe_material(recipe_material_id):
    start_time = time.time()
    material = RecipeMaterial.query.get(recipe_material_id)
    if not material:
        return jsonify({"message": "Recipe material not found"}), 404

    data = request.get_json()
    material.material_id = data.get("material_id", material.material_id)
    material.set_point = data.get("set_point", material.set_point)

    db.session.commit()
    end_time = time.time()
    return jsonify({"message": "Recipe material updated successfully", "execution_time_ms": round((end_time - start_time) * 1000, 2)})

@recipe_bp.route("/recipe_materials/<int:recipe_id>", methods=["GET"])
def get_recipe_materials_by_recipe_id(recipe_id):
    start_time = time.time()
    materials = RecipeMaterial.query.filter_by(recipe_id=recipe_id).all()
    if not materials:
        return jsonify({"message": "No materials found for this recipe"}), 404

    result = [
        {
            "recipe_material_id": mat.recipe_material_id,
            "recipe_id": mat.recipe_id,
            "material_id": mat.material_id,
            "set_point": str(mat.set_point) if mat.set_point is not None else None,
            "actual": str(mat.actual) if mat.actual is not None else None,
            "margin": str(mat.margin) if mat.margin is not None else None
        }
        for mat in materials
    ]
    end_time = time.time()
    return jsonify({"materials": result, "execution_time_ms": round((end_time - start_time) * 1000, 2)})

@recipe_bp.route("/recipe_materials/<int:recipe_material_id>", methods=["DELETE"])
def delete_recipe_material(recipe_material_id):
    start_time = time.time()
    material = RecipeMaterial.query.get(recipe_material_id)
    if not material:
        return jsonify({"message": "Recipe material not found"}), 404

    db.session.delete(material)
    db.session.commit()
    end_time = time.time()
    return jsonify({"message": "Recipe material deleted successfully", "execution_time_ms": round((end_time - start_time) * 1000, 2)})

